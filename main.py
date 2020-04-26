# Libraries for:
# controlling the webbrowser
from selenium import webdriver
# controlling actions in webbrowser(show listed items)
from selenium.webdriver.common.action_chains import ActionChains
# hiding the password
from getpass import getpass
# waiting for the webpages to show up
import time
# processing the data from the site
import openpyxl
from openpyxl import Workbook
# making a path for the files
import os


# Select the driver for the browser
browser = webdriver.Opera(executable_path="operadriver.exe")
# Actions to 'move' a virtual mouse
action = ActionChains(browser)

# Get to the neptun website
browser.get('https://neptun11.uni-pannon.hu/hallgato/login.aspx')

# Search for all the elemnt needed for logging in
username = browser.find_element_by_id('user')
password = browser.find_element_by_id('pwd')
loginBtn = browser.find_element_by_id('btnSubmit')

# Ask the user for their neptun datas !NOT SAVED! and then pass
username.send_keys(input("Kérem a neptun azonosítót: "))
password.send_keys(getpass("Kérem a neptun jelszót: "))
loginBtn.click()

# Wait for the neptun to log in
time.sleep(4)

# This lines search for the courses site
targyakMenu = browser.find_element_by_id('mb1_Targyak')
action.move_to_element(targyakMenu).perform()
targyfelvetel = browser.find_element_by_id('mb1_Targyak_Targyfelvetel')
targyfelvetel.click()

# This checks in a radio buttom for the necessary courses
mintaTantargyakRadio = browser.find_element_by_id('upFilter_rbtnSubjectType_0')
mintaTantargyakRadio.click()
time.sleep(0.5)
listazasButton = browser.find_element_by_id('upFilter_expandedsearchbutton')
listazasButton.click()

"""
This is downloading the must-have data of the courses. 
Have to do it to all, but don't know the name, 
that's why we download all of the information from the site first.

time.sleep(5)


tesztTargy = browser.find_element_by_xpath("//span[text()='Adatelemzés Pythonban']")
tesztTargy.click()

time.sleep(3)

browser.find_element_by_id('Addsubject_course1_gridCourses_imexcel').click()
"""

# For loading all the courses avaible
time.sleep(5)

# Downloads all the data of the courses: named export.xlsx
browser.find_element_by_id('h_addsubjects_gridSubjects_imexcel').click()

# Wait for download the file
path = 'C:\\Users\\soma\\Downloads\export.xlsx'

# Load - wait for it
time.sleep(5)
# - ing

# Open the data xml
book = openpyxl.load_workbook(path)
sheet = book.active

# This will contain all the courses name
coursesName = []

# Here start the names of the courses
column = 'A'
counter = 2

# While there is data in the A cells where are the courses
# Then add them to the list
while True:
    actualCell = sheet[column + str(counter)].value
    if(actualCell == None):
        break
    coursesName.append(actualCell)
    counter = counter + 1

#os.remove(path)

# Just make an xml for all the times
workbook = Workbook()
workbook.save('collected.xlsx')

# Will be needed for the colums, incremented after courses
collectedCounter = 1

# This opens all the courses by it's name, then save them
for course in coursesName:
    courseString = "//span[text()='" + course + "']"
    tempCourse = browser.find_element_by_xpath(courseString)
    tempCourse.click()

    time.sleep(5)

    browser.find_element_by_id('Addsubject_course1_gridCourses_imexcel').click()

    time.sleep(1)

    # Here i set the path of collection of times - the final one
    path = "C:\\Users\\soma\\Downloads\\export (1).xlsx"

    # Open here
    exported = openpyxl.load_workbook(path)
    sheet = exported.active

    # This will contain all the times in the course
    times = []

    # This is where the times start
    column = 'F'
    counter = 2

    # Puts all the times of the courses
    while True:
        actualCell = sheet[column + str(counter)].value
        if(actualCell == None):
            break
        times.append(actualCell)
        counter = counter + 1

    typeO = []

    # The courses type will be put in the typeO list
    column = "B"
    counter = 2

    while True:
        actualCell = sheet[column + str(counter)].value
        if(actualCell == None):
            break
        typeO.append(actualCell)
        counter = counter + 1

    # Remove the excel file, for better disk usage
    os.remove(path)

    # Open the collection, for putting everything from the times list into it
    collected = openpyxl.load_workbook('collected.xlsx')
    sheet = collected.active

    rows = 1

    sheet.cell(row = rows, column = collectedCounter).value = str(course)
    collected.save('collected.xlsx')

    rows += 1

    for i in range(0, len(times)):
        if(times[i] != ""):
            actuallCell = sheet.cell(row = rows, column = collectedCounter).value = times[i] + "?" + typeO[i]
            
        # Save after every insertion
        collected.save('collected.xlsx')
        rows = rows + 1

    # You can't be sure enough
    collected.save('collected.xlsx')

    # Next courses' times go to other column
    collectedCounter = collectedCounter + 1

    # Just close the websites pop up window
    browser.find_element_by_xpath('//*[@title="close"]').click()

path = "C:\\Users\\soma\\Downloads\\export.xlsx"
os.remove(path)

collected = openpyxl.load_workbook('collected.xlsx')
sheet = collected.active

f = open("allCourse.txt", "w")

columnCounter = 1
for i in range(0, len(coursesName)):
    counter = 1

    # Puts all the times of the courses
    while True:
        actualCell = sheet.cell(row = counter, column = columnCounter).value
        if(actualCell != None):
            f.write(actualCell + "\n")
        else:
            break
        counter = counter + 1

    f.write("*\n")

    columnCounter += 1
    
f.close()

